from cobra.io import read_sbml_model
import openpyxl
import glob

excel = openpyxl.Workbook()
ws = excel.active
ws.title = "configuration"
configuration = excel['configuration']
configuration['A1'] = 'Model name'
configuration['B1'] = 'patt'
configuration['C1'] = 'model_file'

excel.create_sheet('states')
states = excel['states']
states['B1'] = 'IC'
states['C1'] = 'ODE (xdot)'
states['D1'] = 'Plot legend'

excel.create_sheet('dynamic_constraints')
dynamic_constraints = excel['dynamic_constraints']
dynamic_constraints['A1'] = 'variable'
dynamic_constraints['B1'] = 'formula'
dynamic_constraints['C1'] = 'reaction_id'
dynamic_constraints['D1'] = 'type'

excel.create_sheet('GSM_constraints')
GSM_constraints = excel['GSM_constraints']
GSM_constraints['A1'] = 'rxn_id'
GSM_constraints['B1'] = 'rxn_name'
GSM_constraints['C1'] = 'lower_bound_dft'
GSM_constraints['D1'] = 'lower_bound_modef'
GSM_constraints['E1'] = 'upper_bound_dft'
GSM_constraints['F1'] = 'upper_bound_modef'

excel.create_sheet('parameters')
parameters = excel['parameters']
parameters['B1'] = 'value'
parameters['C1'] = 'is_opt'

excel.create_sheet('plot')
plot = excel['plot']
plot['A1'] = 'title'
plot['A2'] = 'x axis label'
plot['A3'] = 'y axis label'
plot['A4'] = 'time interval (begin)'
plot['A5'] = 'time interval (end)'

excel.create_sheet('fluxes')
fluxes = excel['fluxes']
fluxes['A1'] = 'time'

x = glob.glob('*.xml')
cobra_model = read_sbml_model(x[0])
reactions = [cobra_model.reactions[reaction].id for reaction in range(len(cobra_model.reactions))]

for line in range(len(reactions)):
	excel_line = str(line + 2)
	r_id = reactions[line]
	GSM_constraints['A'+excel_line] = r_id
	GSM_constraints['B'+excel_line] = cobra_model.reactions.get_by_id(r_id).name
	GSM_constraints['C'+excel_line] = cobra_model.reactions.get_by_id(r_id).lower_bound
	GSM_constraints['E'+excel_line] = cobra_model.reactions.get_by_id(r_id).upper_bound
	fluxes['A'+excel_line] = r_id

excel.save('dfba.xlsx')
