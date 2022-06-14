from cobra.io import read_sbml_model
import openpyxl
import sympy as sym
import numpy as np
from scipy import integrate
import matplotlib.pyplot as plt
from cobra.flux_analysis import pfba
from openpyxl.utils import get_column_letter
import glob
sym.init_printing()


def modification(lst, lb=True):
	"""
	Function that alters lower or upper bounds of reactions
	:param lst: list
	:param lb: bool
	:return:
	"""
	for val in range(len(lst)):
		if lst[val] is not None and lb is True:
			cobra_model.reactions.get_by_id(GSM_constraints['A' + str(val + 2)].value).lower_bound = lst[val]
		elif lst[val] is not None:
			cobra_model.reactions.get_by_id(GSM_constraints['A' + str(val + 2)].value).upper_bound = lst[val]


def bound_types(model, typ, reaction_id, line, val=None):
	"""
		Function that alters lower and upper bounds of the reactions in dynamic_constraints
		:param model: model
		:param typ: str
		:param reaction_id: str
		:param line: str
		:param val: int
		:return:
		"""
	if val is None:
		value = eval(dynamic_constraints['B' + str(line)].value)
	if typ.lower() == 'equality':
		model.reactions.get_by_id(reaction_id).lower_bound = value
		model.reactions.get_by_id(reaction_id).upper_bound = value
	elif typ.lower() == 'upper':
		model.reactions.get_by_id(reaction_id).upper_bound = value
	elif typ.lower() == 'lower':
		model.reactions.get_by_id(reaction_id).lower_bound = value


excel = openpyxl.load_workbook(filename="dfba.xlsx")
configuration = excel['configuration']
states = excel['states']
dynamic_constraints = excel['dynamic_constraints']
GSM_constraints = excel['GSM_constraints']
parameters = excel['parameters']
plot = excel['plot']
flux = excel['fluxes']

x = glob.glob('*.xml')
cobra_model = read_sbml_model(x[0])
# cobra_model = read_sbml_model(configuration['B2'].value)
lb_mdf = [GSM_constraints['D' + str(lb + 1)].value for lb in range(1, len(GSM_constraints['D']))]
up_mdf = [GSM_constraints['F' + str(lb + 1)].value for lb in range(1, len(GSM_constraints['F']))]
modification(lb_mdf)
modification(up_mdf, False)


# differential equations
def xdot(t, x, cobra_model):
	for s in range(2, len(states['A']) + 1):
		if states['A' + str(s)].value is not None:
			globals()[states['A' + str(s)].value] = x[s - 2]

	for param in range(2, len(parameters['A']) + 1):
		if parameters['A' + str(param)].value is None: break
		globals()[parameters['A' + str(param)].value] = parameters['B' + str(param)].value

	for cons in range(2, len(dynamic_constraints['A']) + 1):
		if dynamic_constraints['B' + str(cons)].value is not None and dynamic_constraints['A' + str(cons)].value is not None:
			if type(dynamic_constraints['B' + str(cons)].value) is float or type(dynamic_constraints['B' + str(cons)].value) is int:
				globals()[dynamic_constraints['A' + str(cons)].value] = dynamic_constraints['B' + str(cons)].value
			else:
				globals()[dynamic_constraints['A' + str(cons)].value] = eval(dynamic_constraints['B' + str(cons)].value)
			bound_types(cobra_model, dynamic_constraints['D' + str(cons)].value, dynamic_constraints['C' + str(cons)].value, cons)

	lst = []
	for sta in range(2, len(states['A']) + 1):
		if states['A' + str(sta)].value is not None:
			globals()[states['A' + str(sta)].value] = eval(states['C' + str(sta)].value)
			lst.append(eval(states['A' + str(sta)].value))

	pfba_solution = pfba(cobra_model)

	column = 1
	while True:
		col = get_column_letter(column)
		if flux[col + str(1)].value is None:
			break
		else:
			column += 1
	for line in range(len(flux['A'])):
		if line + 1 == 1: flux[col + str(1)] = t
		if line + 1 != 1: flux[col + str(line + 1)] = pfba_solution.fluxes[flux['A' + str(line + 1)].value]

	return np.array(lst)


t_span = (0, 100)
t_eval = np.linspace(t_span[0], t_span[1], 1000)
y0 = []
for valu in range(2, len(states['B']) + 1):
	if states['B' + str(valu)].value is not None: y0.append(states['B' + str(valu)].value)

sol = integrate.solve_ivp(xdot, t_span, y0, method='LSODA', t_eval=t_eval, args=[cobra_model])

legend_names = []
for var in range(len(sol.y)):
	plt.plot(t_eval, sol.y[var])
	legend_names.append(states['D' + str(var+2)].value)

if plot['B1'].value is not None:
	plt.title(plot['B1'].value)
plt.xlabel(plot['B2'].value)
plt.ylabel(plot['B3'].value)
plt.legend(legend_names)
plt.savefig('plot.pdf')
excel.save('dfba.xlsx')
